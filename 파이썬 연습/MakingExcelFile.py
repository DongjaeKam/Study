line_cnt = 13
from openpyxl import load_workbook
import urllib.request
load_wb = load_workbook("D:\DP_DATA\data.xlsx")
load_ws = load_wb['_itemdownload']
for i in range(0,10):
    for j in range(0,line_cnt):
        url = load_ws.cell(3+j, 24+i).value
        print(url)
        num = i+(j+1)*10
        filename = f"D:\DP_IMAGE\{num}.jpg"
        urllib.request.urlretrieve(url, filename)
        
load_naver = load_wb['naver']
line_cnt = 13

for i in range(0,line_cnt) :
    load_naver.cell(5+i,8).value = str((i+1)*10)+".jpg"
    tmp = str((i+1)*10+1)+".jpg"
    for j in range(2,10):
        tmp+=","+str((i+1)*10+j)+".jpg"
    
    load_naver.cell(5+i,9).value=tmp
    
load_wb.save("D:\DP_DATA\data.xlsx")
